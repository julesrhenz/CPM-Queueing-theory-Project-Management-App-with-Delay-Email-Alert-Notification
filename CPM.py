import tkinter as tk
import hashlib
import math
import smtplib
from email.message import EmailMessage
import re
import os
from PIL import Image, ImageTk
import psycopg2 
from psycopg2 import sql 
import pandas as pd
import matplotlib.pyplot as plt
import traceback
import tempfile
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import datetime, date, timedelta
from openpyxl.drawing.image import Image as XLImage
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tkcalendar import DateEntry
from collections import deque
from collections import defaultdict
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587

SENDER_EMAIL = "julesrhenz.35@gmail.com"
SENDER_APP_PASSWORD = "nxno tyuh dspf hvin"  # NOT your normal Gmail password

def send_email(to_email: str, subject: str, body: str):
        """Send a plain-text email via Gmail SMTP (App Password)."""
        msg = EmailMessage()
        msg["From"] = SENDER_EMAIL
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.set_content(body)

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_APP_PASSWORD)
            server.send_message(msg)

def fetch_delayed_tasks(table_name="tasks"):
    conn = psycopg2.connect(**DB_PARAMS)
    cur = conn.cursor()
    cur.execute(f"""
        SELECT ID, PROJECT_NAME, STATUS, TARGET_DELIVERY
        FROM {table_name}
        WHERE COALESCE(status, '') <> 'Completed'
        AND TARGET_DELIVERY IS NOT NULL
        AND TARGET_DELIVERY < CURRENT_DATE
        ORDER BY TARGET_DELIVERY ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def should_send_alert_today(task_id: int, alert_type="DELAYED") -> bool:
    """Anti-spam: send only once per day per task per alert type."""
    conn = psycopg2.connect(**DB_PARAMS)
    cur = conn.cursor()
    cur.execute("""
        SELECT last_sent
        FROM alert_log
        WHERE task_id=%s AND alert_type=%s
    """, (task_id, alert_type))
    row = cur.fetchone()
    cur.close()
    conn.close()

    if row is None:
        return True

    last_sent = row[0]
    return last_sent.date() < date.today()


def mark_alert_sent(task_id: int, alert_type="DELAYED"):
    conn = psycopg2.connect(**DB_PARAMS)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO alert_log (task_id, alert_type, last_sent)
        VALUES (%s, %s, NOW())
        ON CONFLICT (task_id, alert_type)
        DO UPDATE SET last_sent = NOW()
    """, (task_id, alert_type))
    conn.commit()
    cur.close()
    conn.close()


def fetch_admin_emails() -> list[str]:
    """
    If you want to notify all Admins, this expects your users table like:
    users(username, password, role)
    where role might be 'admin'/'user' OR 'Admin'/'User'.
    """
    conn = psycopg2.connect(**DB_PARAMS)
    cur = conn.cursor()
    cur.execute("""
        SELECT username
        FROM users
        WHERE role IN ('Admin','admin')
        AND username IS NOT NULL
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [r[0] for r in rows]

STATUS_COLORS = {
    "ON TIME": "#c8f7c5",                  # mint green
    "ONGOING": "#c8f7c5",                  # mint green
    "DELAY": "#ffb3b3",                    # light red
    "DELIVERED": "#90ee90",                # green
    "COMPLETED": "#b2ebf2",                # cyan
    "SUBJECT FOR MODIFICATION": "#e6ccff", # light violet
    "ON-HOLD": "#fff3b0",                  # yellow
}
# ---------------- DATABASE ----------------
DB_PARAMS = {
    'dbname': 'postgres',
    'user': 'postgres',
    'password': '1234',
    'host': 'localhost',
    'port': '5433'
}

def create_db():
    conn = psycopg2.connect(**DB_PARAMS)
    cur = conn.cursor()

    # Users table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL
        )
    """)

    # Default tasks table (your current one)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            ID SERIAL PRIMARY KEY,
            CUSTOMER TEXT,
            PROJECT_NAME TEXT NOT NULL,
            DATE_OF_PO DATE NOT NULL,
            LEADTIME REAL NOT NULL,
            STATUS TEXT,
            TARGET_DELIVERY DATE NULL,
            DEPENDENCIES TEXT NULL,
            LAMBDA_VAL REAL NOT NULL,
            MU_VAL REAL NOT NULL,
            NOS_VAL INTEGER NOT NULL,
            WORK_TYPE TEXT NOT NULL,
            REMARKS TEXT
        )
    """)

    # Registry for per-file tables (Option B)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS imported_files (
            id SERIAL PRIMARY KEY,
            display_name TEXT NOT NULL,
            table_name TEXT NOT NULL UNIQUE,
            file_path TEXT,
            imported_at TIMESTAMP DEFAULT NOW()
        )
    """)

    conn.commit()
    cur.close()
    conn.close()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def register_user(username, password, role):
    try:
        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s)",
            (username, hash_password(password), role.lower())
        )
        conn.commit()
        cur.close()
        conn.close()
        return True, "User registered successfully"
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        cur.close()
        conn.close()
        return False, "Username already exists"
    except Exception as e:
        return False, str(e)

def verify_login(username, password):
    conn = psycopg2.connect(**DB_PARAMS)
    cur = conn.cursor()
    cur.execute(
        "SELECT password_hash, role FROM users WHERE username=%s",
        (username,)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return False, "User not found", None
    stored_hash, role = row
    if stored_hash == hash_password(password):
        return True, "Login successful", role
    return False, "Incorrect password", None

# ---------------- APP CLASS ----------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Project Management App")
        self.geometry("1300x700")
        self.resizable(True, True)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background="#f0f0f0")
        style.configure("Header.TLabel", font=("Arial", 16, "bold"), foreground="#333")
        style.configure("TButton", font=("Arial", 11), padding=5)
        style.map("TButton",
                  background=[('active','#0052cc')],
                  foreground=[('active','white')])

        container = ttk.Frame(self)
        container.pack(fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        self.frames = {}
        for F in (LoginPage, DashboardPage):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame(LoginPage)

    def show_frame(self, page_class):
        frame = self.frames[page_class]
        frame.tkraise()

EMAIL_REGEX = r"^[\w\.-]+@[\w\.-]+\.\w+$"

class LoginPage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        ttk.Label(
            self,
            text="Project Management Login System",
            style="Header.TLabel"
        ).pack(pady=(30, 20))

        # EMAIL
        ttk.Label(self, text="Email:").pack(pady=(10, 5))
        self.email_entry = ttk.Entry(self, width=30)
        self.email_entry.pack()

        # PASSWORD
        ttk.Label(self, text="Password:").pack(pady=(10, 5))
        self.pass_entry = ttk.Entry(self, width=30, show="*")
        self.pass_entry.pack()

        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=30)

        ttk.Button(
            btn_frame, text="Log In", width=15, command=self.login
        ).pack(side="left", padx=15)

        ttk.Button(
            btn_frame, text="Sign Up", width=15, command=self.signup
        ).pack(side="left", padx=15)

    # ---------------- LOGIN ----------------
    def login(self):
        email = self.email_entry.get().strip().lower()
        password = self.pass_entry.get().strip()

        if not re.match(EMAIL_REGEX, email):
            messagebox.showerror("Error", "Please enter a valid email address")
            return

        ok, msg, role = verify_login(email, password)

        if ok:
            role = role.capitalize() 
            messagebox.showinfo("Success", f"{msg}! Role: {role}")
            dash_page = self.controller.frames[DashboardPage]
            dash_page.set_user(email, role)
            self.controller.show_frame(DashboardPage)

            if not getattr(dash_page, "_alerts_started", False):
                dash_page._alerts_started = True
                dash_page.start_alert_loop()
        else:
            messagebox.showerror("Error", msg)

    # ---------------- SIGN UP ----------------
    def signup(self):
        def perform_signup():
            email = su_email.get().strip().lower()   # email can be lowercase
            pwd = su_pass.get().strip()
            role = su_role.get().strip()              # KEEP CASE

            if not re.match(EMAIL_REGEX, email):
                messagebox.showerror("Error", "Invalid email format")
                return

            if role not in ("Admin", "User"):
                messagebox.showerror(
                    "Error",
                    "Role must be exactly 'Admin' or 'User' (case-sensitive)"
                )
                return

            ok, msg = register_user(email, pwd, role)
            if ok:
                messagebox.showinfo("Success", msg)
                signup_win.destroy()
            else:
                messagebox.showerror("Error", msg)

        signup_win = tk.Toplevel(self)
        signup_win.title("Sign Up")
        signup_win.geometry("350x260")
        signup_win.resizable(False, False)

        ttk.Label(signup_win, text="Email:").pack(pady=5)
        su_email = ttk.Entry(signup_win, width=30)
        su_email.pack()

        ttk.Label(signup_win, text="Password:").pack(pady=5)
        su_pass = ttk.Entry(signup_win, width=30, show="*")
        su_pass.pack()

        ttk.Label(signup_win, text="Role (Admin/User):").pack(pady=5)
        su_role = ttk.Entry(signup_win, width=30)
        su_role.pack()

        ttk.Button(
            signup_win, text="Sign Up", command=perform_signup
        ).pack(pady=15)

# ---------------- DASHBOARD PAGE ----------------
class DashboardPage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.current_user = ""
        self.current_role = ""
        # active dataset table
        self.current_tasks_table = "tasks"

        # dataset selector state
        self.dataset_var = tk.StringVar(value="Default (tasks)")
        self.dataset_map = {"Default (tasks)": "tasks"}
        create_db()

        # Layout
        self.sidebar = ttk.Frame(self)
        self.sidebar.pack(side="left", fill="y", padx=5, pady=5)
        self.sidebar.pack_propagate(False)
        self.sidebar.configure(width=260)  # ✅ prevents sidebar from shrinking
        self.main_area = ttk.Frame(self)
        self.main_area.pack(side="right", fill="both", expand=True, padx=5, pady=5)

        # Build the left panel sections (TOP LEFT)
        self.build_user_session_info(self.sidebar)
        self.build_kpi_cards(self.sidebar)
        self.build_activity_log(self.sidebar)

        self.notebook = ttk.Notebook(self.main_area)
        self.notebook.pack(fill="both", expand=True)

        self.tab_dashboard = ttk.Frame(self.notebook)
        self.tab_tasks = ttk.Frame(self.notebook)
        self.tab_cpm = ttk.Frame(self.notebook)
        self.tab_queueing = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_dashboard, text="Dashboard")
        self.notebook.add(self.tab_tasks, text="Projects")
        self.notebook.add(self.tab_cpm, text="CPM")
        self.notebook.add(self.tab_queueing, text="Queueing")
        self._last_summary_rows = []

        # ---------------- QUEUEING TAB ----------------
        self.tab_queueing.grid_columnconfigure(0, weight=1)
        self.tab_queueing.grid_rowconfigure(0, weight=0)
        self.tab_queueing.grid_rowconfigure(1, weight=0)
        self.tab_queueing.grid_rowconfigure(2, weight=2)
        self.tab_queueing.grid_rowconfigure(3, weight=2, minsize=0)

        # Top controls
        queue_controls = ttk.Frame(self.tab_queueing)
        queue_controls.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 10))
        queue_controls.grid_columnconfigure(0, weight=1)  # optional

        ttk.Button(queue_controls, text="CALCULATE QUEUEING", command=self.calculate_queueing).pack(side="left", padx=5)
        ttk.Button(queue_controls, text="SHOW HEATMAP", command=self.show_heatmap).pack(side="left", padx=5)
        ttk.Button(queue_controls, text="EXPORT QUEUEING", command=self.export_queueing_excel).pack(side="right", padx=10)

        self._add_dynamic_banner_pack(queue_controls, "D:/Downloads/vtech.png", height=28)

        # optional: ensure it's on top
        queue_controls.tkraise()

        # ---- Summary (WORK TYPE) ----
        summary_frame = ttk.LabelFrame(self.tab_queueing, text="SUMMARY BY WORK TYPE")
        summary_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 8))
        summary_frame.grid_rowconfigure(0, weight=1)
        summary_frame.grid_columnconfigure(0, weight=1)

        sum_cols = ("WORK TYPE", "COUNT", "TOTAL_LAMBDA", "AVG_MU", "TOTAL_SERVERS", "AVG_RHO", "MAX_RHO", "FLAG")
        # Vertical scrollbar
        sum_y = ttk.Scrollbar(summary_frame, orient="vertical")
        sum_y.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar  ✅ ADD THIS
        sum_x = ttk.Scrollbar(summary_frame, orient="horizontal")
        sum_x.grid(row=1, column=0, sticky="ew")

        self.queue_summary_tree = ttk.Treeview(summary_frame, columns=sum_cols, show="headings", yscrollcommand=sum_y.set,
        xscrollcommand=sum_x.set)  # REMOVE height=6
        self.queue_summary_tree.grid(row=0, column=0, sticky="nsew")  # nsew (not ew)

        sum_y.config(command=self.queue_summary_tree.yview)
        sum_x.config(command=self.queue_summary_tree.xview)

        for c in sum_cols:
            self.queue_summary_tree.heading(c, text=c)
            self.queue_summary_tree.column(c, width=130, anchor="center")
        self.queue_summary_tree.column("WORK TYPE", width=200, anchor="w")
        self.queue_summary_tree.column("FLAG", width=220, anchor="w")

        # ---- Detail Table ----
        detail_frame = ttk.LabelFrame(self.tab_queueing, text="ALL PROJECTS")
        detail_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 8))
        detail_frame.grid_rowconfigure(0, weight=1)
        detail_frame.grid_columnconfigure(0, weight=1)

        queue_y = ttk.Scrollbar(detail_frame, orient="vertical")
        queue_y.grid(row=0, column=1, sticky="ns")

        queue_x = ttk.Scrollbar(detail_frame, orient="horizontal")
        queue_x.grid(row=1, column=0, sticky="ew")

        queue_cols = (
            "ID","CUSTOMER","PROJECT NAME","WORK TYPE","LAMBDA","MU","SERVERS",
            "RHO","WQ","W","LQ","L","P0","NOTE"
        )

        self.queue_tree = ttk.Treeview(
            detail_frame,
            columns=queue_cols,
            show="headings",
            yscrollcommand=queue_y.set,
            xscrollcommand=queue_x.set
        )
        self.queue_tree.grid(row=0, column=0, sticky="nsew")

        queue_y.config(command=self.queue_tree.yview)
        queue_x.config(command=self.queue_tree.xview)

        for c in queue_cols:
            self.queue_tree.heading(c, text=c)
            self.queue_tree.column(c, width=120, anchor="center")

        self.queue_tree.column("CUSTOMER", width=220, anchor="w")
        self.queue_tree.column("PROJECT NAME", width=220, anchor="w")
        self.queue_tree.column("WORK TYPE", width=180, anchor="center")
        self.queue_tree.column("NOTE", width=360, anchor="w")

        # ---- Row coloring (unstable / warning / ok) ----
        self.queue_tree.tag_configure("unstable", background="#ffb3b3")  # light red
        self.queue_tree.tag_configure("warning", background="#fff3b0")   # light yellow
        self.queue_tree.tag_configure("ok", background="#c8f7c5")        # mint green

        self.queue_summary_tree.tag_configure("unstable", background="#ffb3b3")
        self.queue_summary_tree.tag_configure("warning", background="#fff3b0")
        self.queue_summary_tree.tag_configure("ok", background="#c8f7c5")

        # ---- Heatmap area (matplotlib) ----
        self.heatmap_frame = ttk.LabelFrame(self.tab_queueing, text="UTILIZATION HEATMAP (WORK TYPE x METRICS)")
        heatmap_frame = self.heatmap_frame
        heatmap_frame.grid(row=3, column=0, sticky="nsew", padx=10, pady=(0, 10))
        heatmap_frame.grid_rowconfigure(0, weight=1)
        heatmap_frame.grid_columnconfigure(0, weight=1)

        # IMPORTANT: create the Figure ONCE
        self.queue_heatmap_fig = Figure(figsize=(10, 2.8), dpi=100)

        # Create the canvas ONCE
        self.queue_heatmap_canvas = FigureCanvasTkAgg(
            self.queue_heatmap_fig,
            master=heatmap_frame
        )

        canvas_widget = self.queue_heatmap_canvas.get_tk_widget()
        canvas_widget.grid(row=0, column=0, sticky="nsew")

        # ---------------- DASHBOARD LABEL ----------------
        self._build_dashboard_ui()
        self.refresh_dashboard()
        self._build_projects_ui()
        self._build_cpm_ui()

    def build_user_session_info(self, parent):
        user_frame = ttk.LabelFrame(parent, text="User Session", height=90)
        user_frame.pack(fill="x", padx=5, pady=(0, 6))
        user_frame.pack_propagate(False)  # IMPORTANT

        self.user_label = ttk.Label(
            user_frame,
            text="Logged in: --\nRole: --\nLast refresh: --",
            justify="left"
        )
        self.user_label.pack(anchor="w", padx=10, pady=8)

    def build_kpi_cards(self, parent):
        kpi_frame = ttk.LabelFrame(parent, text="Key Metrics", height=120)
        kpi_frame.pack(fill="x", padx=5, pady=(0, 6))
        kpi_frame.pack_propagate(False)  # IMPORTANT

        self.lbl_avg_lead = ttk.Label(kpi_frame, text="Avg Lead Time: -- days")
        self.lbl_on_time = ttk.Label(kpi_frame, text="On-Time Rate: --%")
        self.lbl_avg_delay = ttk.Label(kpi_frame, text="Avg Delay: -- days")
        self.lbl_critical = ttk.Label(kpi_frame, text="Critical Projects: --")

        for lbl in (
            self.lbl_avg_lead,
            self.lbl_on_time,
            self.lbl_avg_delay,
            self.lbl_critical
        ):
            lbl.pack(anchor="w", pady=1)

    def build_activity_log(self, parent):
        activity_frame = ttk.LabelFrame(parent, text="Activity Log")
        activity_frame.pack(fill="both", expand=True, padx=5)
        activity_frame.pack_propagate(False)

        # Activity list
        self.activity_list = tk.Listbox(
            activity_frame,
            height=8,
            activestyle="none"
        )
        self.activity_list.pack(fill="both", expand=True, padx=4, pady=(4, 6))

        self.activity_list.insert("end", "No recent activity")

        # ---- Logout button (BOTTOM) ----
        logout_btn = ttk.Button(
            activity_frame,
            text="Logout",
            command=self.logout_user
        )
        logout_btn.pack(fill="x", padx=6, pady=(0, 6))


    def refresh_activity_log(self):
        self.activity_list.delete(0, "end")

        sql = """
            SELECT project_name, status, target_delivery
            FROM tasks
            ORDER BY id DESC
            LIMIT 10
        """

        from datetime import date

        today = date.today()

        with self._db() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                rows = cur.fetchall()

        for name, status, tdate in rows:
            status_u = (status or "").upper()

            if status_u == "DELAY":
                msg = f"⚠ {name} is now DELAYED"
            elif status_u.startswith("COMPLETED"):
                msg = f"✔ {name} completed"
            elif tdate and (tdate - today).days <= 3:
                msg = f"⏰ {name} due soon"
            else:
                msg = f"ℹ {name} updated"

            self.activity_list.insert("end", msg)

        if not rows:
            self.activity_list.insert("end", "No recent activity")

    def _build_dashboard_ui(self):
        tab = self.tab_dashboard
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(0, weight=0)  # KPIs
        tab.grid_rowconfigure(1, weight=1)  # alerts + actions
        tab.grid_rowconfigure(2, weight=2)  # charts
        tab.grid_rowconfigure(3, weight=2)  # nearest deadlines

        # ---- Row 0: KPI cards ----
        kpi_frame = ttk.Frame(tab)
        kpi_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 8))
        for i in range(6):
            kpi_frame.grid_columnconfigure(i, weight=1, uniform="kpi")

        self.kpi_vars = {
            "TOTAL": tk.StringVar(value="0"),
            "ONGOING": tk.StringVar(value="0"),
            "ON-HOLD": tk.StringVar(value="0"),
            "DELAY": tk.StringVar(value="0"),
            "ON TIME": tk.StringVar(value="0"),
            "COMPLETED": tk.StringVar(value="0"),
        }

        def kpi_card(parent, col, title, var):
            card = ttk.LabelFrame(parent, text=title)
            card.grid(row=0, column=col, sticky="nsew", padx=5)
            lbl = ttk.Label(card, textvariable=var, font=("Segoe UI", 16, "bold"))
            lbl.pack(padx=10, pady=8)
            return card

        kpi_card(kpi_frame, 0, "TOTAL PROJECTS", self.kpi_vars["TOTAL"])
        kpi_card(kpi_frame, 1, "ONGOING", self.kpi_vars["ONGOING"])
        kpi_card(kpi_frame, 2, "ON-HOLD", self.kpi_vars["ON-HOLD"])
        kpi_card(kpi_frame, 3, "DELAY", self.kpi_vars["DELAY"])
        kpi_card(kpi_frame, 4, "ON TIME", self.kpi_vars["ON TIME"])
        kpi_card(kpi_frame, 5, "COMPLETED", self.kpi_vars["COMPLETED"])

        # ---- Row 1: Alerts + Quick actions ----
        mid = ttk.Frame(tab)
        mid.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 8))
        mid.grid_columnconfigure(0, weight=3)
        mid.grid_columnconfigure(1, weight=1)
        mid.grid_rowconfigure(0, weight=1)

        # Alerts
        alerts_frame = ttk.LabelFrame(mid, text="ALERTS / RISKS")
        alerts_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        alerts_frame.grid_columnconfigure(0, weight=1)
        alerts_frame.grid_rowconfigure(0, weight=1)

        alert_cols = ("TYPE", "ID", "CUSTOMER", "PROJECT NAME", "DETAILS")
        self.alert_tree = ttk.Treeview(alerts_frame, columns=alert_cols, show="headings")
        self.alert_tree.grid(row=0, column=0, sticky="nsew")

        a_y = ttk.Scrollbar(alerts_frame, orient="vertical", command=self.alert_tree.yview)
        a_y.grid(row=0, column=1, sticky="ns")
        self.alert_tree.configure(yscrollcommand=a_y.set)

        for c in alert_cols:
            self.alert_tree.heading(c, text=c)
            self.alert_tree.column(c, width=150, anchor="w", stretch=True)
        self.alert_tree.column("TYPE", width=140    , anchor="w", stretch=True)
        self.alert_tree.column("ID", width=60, anchor="center", stretch=True)
        self.alert_tree.column("DETAILS", width=260, anchor="w", stretch=True)

        # click alert → go to tab + select
        self.alert_tree.bind("<Double-1>", self._on_alert_double_click)

        # Quick actions
        actions = ttk.LabelFrame(mid, text="QUICK ACTIONS")
        actions.grid(row=0, column=1, sticky="nsew", padx=10, pady=(0, 8))
        actions.grid_rowconfigure(0, weight=1)
        actions.grid_columnconfigure(0, weight=1)

        # one inner frame that fills the whole quick actions area
        btn_box = ttk.Frame(actions)
        btn_box.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        btn_box.grid_columnconfigure(0, weight=1)

        btns = [
            ("Refresh Dashboard", self.refresh_dashboard),
            ("Load Dataset", self.load_selected_dataset),
            ("Import Excel", self.import_excel),
            ("Export All Tasks", self.export_tasks_excel),
        ]

        for r, (text, cmd) in enumerate(btns):
            btn_box.grid_rowconfigure(r, weight=1)  # optional: distributes vertical space evenly
            ttk.Button(btn_box, text=text, command=cmd).grid(
                row=r, column=0,
                sticky="ew",          # ✅ button stretches horizontally
                padx=6, pady=6,
                ipady=6               # ✅ button height
            )

        dataset_frame = ttk.Frame(mid)
        dataset_frame.grid(row=1, column=1, sticky="ew", padx=10, pady=(4, 10))

        dataset_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(dataset_frame, text="Dataset:").grid(
            row=0, column=0, sticky="w", padx=(0, 6)
        )

        self.dataset_combo_one = ttk.Combobox(
            dataset_frame,
            textvariable=self.dataset_var,
            values=list(self.dataset_map.keys()),
            state="readonly"
        )
        self.dataset_combo_one.grid(row=0, column=1, sticky="ew")

        # ---- Row 2: Charts ----
        charts = ttk.Frame(tab)
        charts.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 8))
        charts.grid_columnconfigure(0, weight=1)
        charts.grid_columnconfigure(1, weight=1)
        charts.grid_rowconfigure(0, weight=1)

        self.status_chart_frame = ttk.LabelFrame(charts, text="STATUS DISTRIBUTION")
        self.status_chart_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self.status_chart_frame.grid_rowconfigure(0, weight=1)
        self.status_chart_frame.grid_columnconfigure(0, weight=1)

        self.lead_chart_frame = ttk.LabelFrame(charts, text="LEADTIME TREND (LATEST 20)")
        self.lead_chart_frame.grid(row=0, column=1, sticky="nsew")
        self.lead_chart_frame.grid_rowconfigure(0, weight=1)
        self.lead_chart_frame.grid_columnconfigure(0, weight=1)

        # create figures once
        self.status_fig = Figure(figsize=(4, 2.4), dpi=100)
        self.status_ax = self.status_fig.add_subplot(111)
        self.status_canvas = FigureCanvasTkAgg(self.status_fig, master=self.status_chart_frame)
        self.status_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")

        self.lead_fig = Figure(figsize=(4, 2.4), dpi=100)
        self.lead_ax = self.lead_fig.add_subplot(111)
        self.lead_canvas = FigureCanvasTkAgg(self.lead_fig, master=self.lead_chart_frame)
        self.lead_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")

        # ---- Row 3: Nearest deadlines ----
        near = ttk.LabelFrame(tab, text="NEAREST DEADLINES (DOUBLE-CLICK TO OPEN PROJECT)")
        near.grid(row=3, column=0, sticky="nsew", padx=10, pady=(0, 10))
        near.grid_columnconfigure(0, weight=1)
        near.grid_rowconfigure(0, weight=1)

        near_cols = ("ID", "CUSTOMER", "PROJECT NAME", "STATUS", "TARGET DELIVERY", "LEADTIME")
        self.near_tree = ttk.Treeview(near, columns=near_cols, show="headings")
        self.near_tree.grid(row=0, column=0, sticky="nsew")

        n_y = ttk.Scrollbar(near, orient="vertical", command=self.near_tree.yview)
        n_y.grid(row=0, column=1, sticky="ns")
        self.near_tree.configure(yscrollcommand=n_y.set)

        for c in near_cols:
            self.near_tree.heading(c, text=c)
            self.near_tree.column(c, width=140, anchor="w", stretch=True)
        self.near_tree.column("ID", width=60, anchor="center", stretch=True)
        self.near_tree.column("LEADTIME", width=90, anchor="center", stretch=True)

        self.near_tree.bind("<Double-1>", self._on_near_double_click)
    
    def _build_projects_ui(self):
        self.tab_tasks.grid_columnconfigure(0, weight=1)
        self.tab_tasks.grid_rowconfigure(1, weight=1)

        # Row 0 toolbar
        topbar = ttk.Frame(self.tab_tasks)
        topbar.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 6))
        topbar.grid_columnconfigure(4, weight=1)

        ttk.Button(topbar, text="Add Task", command=self.add_task).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(topbar, text="Edit Task", command=self.edit_task).grid(row=0, column=1, padx=(0, 6))
        ttk.Button(topbar, text="Delete Task", command=self.delete_task).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(topbar, text="Refresh", command=self.load_tasks).grid(row=0, column=3, padx=(0, 5))

        self._add_dynamic_banner_grid(
            parent=topbar,
            image_path="D:/Downloads/vtech.png",
            row=0, col=4,
            height=28
        )

        ttk.Button(topbar, text="IMPORT EXCEL", command=self.import_excel).grid(row=0, column=5, padx=(0, 6), sticky="e")
        ttk.Button(topbar, text="EXPORT ALL TASKS", command=self.export_tasks_excel).grid(row=0, column=6, sticky="e")

        # Row 1 table frame
        table_frame = ttk.Frame(self.tab_tasks)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 6))
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        y_scroll = ttk.Scrollbar(table_frame, orient="vertical")
        y_scroll.grid(row=0, column=1, sticky="ns")

        # Row 2 bottom bar
        bottom_bar = ttk.Frame(self.tab_tasks)
        bottom_bar.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 6))

        ttk.Label(bottom_bar, text="Dataset:").pack(side="left", padx=(0, 6))

        self.dataset_combo_two = ttk.Combobox(
            bottom_bar,
            textvariable=self.dataset_var,
            values=list(self.dataset_map.keys()),
            state="readonly",
            width=22
        )
        self.dataset_combo_two.pack(side="left", padx=(0, 8))

        ttk.Button(bottom_bar, text="Load Dataset", command=self.load_selected_dataset).pack(side="left")

        # Row 3 horizontal scrollbar
        x_scroll = ttk.Scrollbar(self.tab_tasks, orient="horizontal")
        x_scroll.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 10))

        cols = ("ID","CUSTOMER","PROJECT NAME","DATE OF PO","LEADTIME","STATUS",
                "TARGET DELIVERY","DEPENDENCIES","LAMBDA","MU","NO. OF SERVER","WORK TYPE","REMARKS")

        self.task_tree = ttk.Treeview(
            table_frame,
            columns=cols,
            show="headings",
            xscrollcommand=x_scroll.set,
            yscrollcommand=y_scroll.set
        )
        self.task_tree.grid(row=0, column=0, sticky="nsew")
        x_scroll.config(command=self.task_tree.xview)
        y_scroll.config(command=self.task_tree.yview)

        for col in cols:
            self.task_tree.heading(col, text=col)

        # widths (keep yours if you like)
        self.task_tree.column("ID", width=60, anchor="center")
        self.task_tree.column("CUSTOMER", width=140, anchor="center")
        self.task_tree.column("PROJECT NAME", width=180, anchor="center")
        self.task_tree.column("DATE OF PO", width=110, anchor="center")
        self.task_tree.column("LEADTIME", width=90, anchor="center")
        self.task_tree.column("STATUS", width=140, anchor="center")
        self.task_tree.column("TARGET DELIVERY", width=130, anchor="center")
        self.task_tree.column("DEPENDENCIES", width=120, anchor="center")
        self.task_tree.column("LAMBDA", width=90, anchor="center")
        self.task_tree.column("MU", width=90, anchor="center")
        self.task_tree.column("NO. OF SERVER", width=90, anchor="center")
        self.task_tree.column("WORK TYPE", width=180, anchor="center")
        self.task_tree.column("REMARKS", width=220, anchor="center")

        for status, color in STATUS_COLORS.items():
            self.task_tree.tag_configure(status, background=color)

        self.refresh_dataset_list()
        self.load_tasks()


    def _build_cpm_ui(self):
        self.tab_cpm.grid_columnconfigure(0, weight=1)
        self.tab_cpm.grid_rowconfigure(0, weight=0)
        self.tab_cpm.grid_rowconfigure(1, weight=1)
        self.tab_cpm.grid_rowconfigure(2, weight=1)
        self.tab_cpm.grid_rowconfigure(3, weight=2)

        cpm_controls = ttk.Frame(self.tab_cpm)
        cpm_controls.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 10))

        ttk.Button(cpm_controls, text="CALCULATE CPM", command=self.show_cpm_table_in_tab).pack(side="left", padx=5)
        ttk.Button(cpm_controls, text="SHOW GANTT CHART", command=self.draw_cpm_gantt).pack(side="left", padx=10)
        ttk.Button(cpm_controls, text="EXPORT CPM", command=self.export_cpm_excel).pack(side="right", padx=10)

        self._add_dynamic_banner_pack(cpm_controls, "D:/Downloads/vtech.png", height=28)

        # CPM results
        cpm_result_frame = ttk.LabelFrame(self.tab_cpm, text="CPM RESULTS")
        cpm_result_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 8))
        cpm_result_frame.grid_rowconfigure(0, weight=1)
        cpm_result_frame.grid_columnconfigure(0, weight=1)

        cpm_cols = ('ID','CUSTOMER','PROJECT NAME','LEADTIME','DEPENDENCIES','ES','EF','LS','LF','SLACK','CRITICAL')
        self.cpm_tree = ttk.Treeview(cpm_result_frame, columns=cpm_cols, show="headings")
        self.cpm_tree.grid(row=0, column=0, sticky="nsew")

        cpm_y = ttk.Scrollbar(cpm_result_frame, orient="vertical")
        cpm_y.grid(row=0, column=1, sticky="ns")
        cpm_x = ttk.Scrollbar(cpm_result_frame, orient="horizontal")
        cpm_x.grid(row=1, column=0, sticky="ew")

        self.cpm_tree.configure(yscrollcommand=cpm_y.set, xscrollcommand=cpm_x.set)
        cpm_y.config(command=self.cpm_tree.yview)
        cpm_x.config(command=self.cpm_tree.xview)

        self.cpm_tree.tag_configure("critical", background="#ffcccc")
        self.cpm_tree.tag_configure("normal", background="#ffffff")

        for c in cpm_cols:
            self.cpm_tree.heading(c, text=c)
            self.cpm_tree.column(c, width=100, anchor="center")
        self.cpm_tree.column("PROJECT NAME", width=220, anchor="w")
        self.cpm_tree.column("DEPENDENCIES", width=180, anchor="center")

        # ALL PROJECTS table inside CPM tab
        cpm_all_frame = ttk.LabelFrame(self.tab_cpm, text="ALL PROJECTS")
        cpm_all_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 8))
        cpm_all_frame.grid_rowconfigure(0, weight=1)
        cpm_all_frame.grid_columnconfigure(0, weight=1)

        all_cols = ("ID","CUSTOMER","PROJECT NAME","DATE OF PO","LEADTIME","STATUS","TARGET DELIVERY",
                    "DEPENDENCIES","LAMBDA","MU","NO. OF SERVER", "WORK TYPE", "REMARKS")
        self.cpm_all_tasks_tree = ttk.Treeview(cpm_all_frame, columns=all_cols, show="headings")
        self.cpm_all_tasks_tree.grid(row=0, column=0, sticky="nsew")

        all_y = ttk.Scrollbar(cpm_all_frame, orient="vertical")
        all_y.grid(row=0, column=1, sticky="ns")
        all_x = ttk.Scrollbar(cpm_all_frame, orient="horizontal")
        all_x.grid(row=1, column=0, sticky="ew")

        self.cpm_all_tasks_tree.configure(yscrollcommand=all_y.set, xscrollcommand=all_x.set)
        all_y.config(command=self.cpm_all_tasks_tree.yview)
        all_x.config(command=self.cpm_all_tasks_tree.xview)

        for c in all_cols:
            self.cpm_all_tasks_tree.heading(c, text=c)
            self.cpm_all_tasks_tree.column(c, width=130, anchor="center")

        # chart frame
        self.cpm_chart_frame = ttk.LabelFrame(self.tab_cpm, text="GANTT CHART")
        self.cpm_chart_frame.grid(row=3, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self.cpm_chart_frame.grid_rowconfigure(0, weight=1)
        self.cpm_chart_frame.grid_columnconfigure(0, weight=1)

    def refresh_dashboard(self):
        try:
            table = getattr(self, "current_tasks_table", "tasks")

            # sidebar
            self.refresh_user_session()
            self.refresh_key_metrics()
            self.refresh_activity_log()

            # main dashboard
            kpis = self._fetch_kpis(table)
            for k, v in kpis.items():
                if k in self.kpi_vars:
                    self.kpi_vars[k].set(str(v))

            self._fill_alerts(table)
            self._draw_status_chart(table)
            self._draw_leadtime_chart(table)
            self._fill_nearest_deadlines(table)

        except Exception as e:
            messagebox.showerror("Dashboard Error", str(e))

    def _db(self):
        return psycopg2.connect(**DB_PARAMS)

    def _fetch_kpis(self, table):
        # counts by status
        sql = f"""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN UPPER(status) = 'ONGOING' THEN 1 ELSE 0 END) AS ongoing,
                SUM(CASE WHEN UPPER(status) = 'ON-HOLD' THEN 1 ELSE 0 END) AS onhold,
                SUM(CASE WHEN UPPER(status) = 'DELAY' THEN 1 ELSE 0 END) AS delay,
                SUM(CASE WHEN UPPER(status) = 'ON TIME' THEN 1 ELSE 0 END) AS ontime,
                SUM(CASE WHEN UPPER(status) LIKE 'COMPLETED%%' THEN 1 ELSE 0 END) AS completed
            FROM {table}
        """
        with self._db() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                row = cur.fetchone()

        return {
            "TOTAL": row[0] or 0,
            "ONGOING": row[1] or 0,
            "ON-HOLD": row[2] or 0,
            "DELAY": row[3] or 0,
            "ON TIME": row[4] or 0,
            "COMPLETED": row[5] or 0,
        }
        
    def _fill_alerts(self, table):
        self.alert_tree.delete(*self.alert_tree.get_children())

        from datetime import date, timedelta

        today = date.today()
        soon = today + timedelta(days=7)
        target_col = "target_delivery"

        # 1) DELAY (include target_delivery for consistent detail)
        sql_delay = f"""
            SELECT id, customer, project_name, status, {target_col}
            FROM {table}
            WHERE UPPER(status) = 'DELAY'
            ORDER BY id
            LIMIT 50
        """

        # 2) DUE SOON (ignore completed robustly)
        sql_due = f"""
            SELECT id, customer, project_name, status, {target_col}
            FROM {table}
            WHERE {target_col} IS NOT NULL
            AND {target_col}::date BETWEEN %s AND %s
            AND UPPER(COALESCE(status,'')) NOT LIKE 'COMPLETED%%'
            ORDER BY {target_col}::date ASC
            LIMIT 50
        """

        with self._db() as conn:
            with conn.cursor() as cur:
                # DELAY
                cur.execute(sql_delay)
                for pid, cust, name, status, tdate in cur.fetchall():
                    self.alert_tree.insert(
                        "", "end",
                        values=("DELAY", pid, cust, name, f"Due: {tdate} | Status: {status}")
                    )

                # DUE SOON
                cur.execute(sql_due, (today, soon))
                for pid, cust, name, status, tdate in cur.fetchall():
                    self.alert_tree.insert(
                        "", "end",
                        values=("DUE SOON", pid, cust, name, f"Due: {tdate} | Status: {status}")
                    )

        # 3) Queue unstable (still different by nature)
        if hasattr(self, "queue_tree"):
            for item_id in self.queue_tree.get_children():
                vals = self.queue_tree.item(item_id, "values")
                try:
                    rho = float(vals[7])
                    if rho >= 1.0:
                        pid, cust, pname = vals[0], vals[1], vals[2]
                        self.alert_tree.insert(
                            "", "end",
                            values=("QUEUE UNSTABLE", pid, cust, pname, f"ρ = {rho:.3f}")
                        )
                except:
                    pass
        
    def _draw_status_chart(self, table):
        self.status_ax.clear()

        sql = f"""
            SELECT UPPER(status) AS s, COUNT(*)
            FROM {table}
            GROUP BY UPPER(status)
            ORDER BY COUNT(*) DESC
        """
        with self._db() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                rows = cur.fetchall()

        labels = [r[0] for r in rows]
        counts = [r[1] for r in rows]

        self.status_ax.bar(labels, counts)
        self.status_ax.set_ylabel("Count")
        self.status_ax.tick_params(axis="x", labelrotation=25)
        self.status_fig.tight_layout()
        self.status_canvas.draw()

    def _draw_leadtime_chart(self, table):
        self.lead_ax.clear()

        # leadtime trend of latest 20 by ID
        sql = f"""
            SELECT id, leadtime
            FROM {table}
            WHERE leadtime IS NOT NULL
            ORDER BY id DESC
            LIMIT 20
        """
        with self._db() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                rows = cur.fetchall()

        rows = list(reversed(rows))
        x = [r[0] for r in rows]
        y = [float(r[1]) for r in rows]

        self.lead_ax.plot(x, y, marker="o")
        self.lead_ax.set_xlabel("Project ID")
        self.lead_ax.set_ylabel("Leadtime")
        self.lead_fig.tight_layout()
        self.lead_canvas.draw()

    def _fill_nearest_deadlines(self, table):
        self.near_tree.delete(*self.near_tree.get_children())

        target_col = 'target_delivery'  # <-- adjust if needed
        sql = f"""
            SELECT id, customer, project_name, status, {target_col}, leadtime
            FROM {table}
            WHERE {target_col} IS NOT NULL
            ORDER BY {target_col}::date ASC
            LIMIT 15
        """
        with self._db() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                for pid, cust, name, status, tdate, lead in cur.fetchall():
                    self.near_tree.insert("", "end", values=(pid, cust, name, status, str(tdate), lead))

    def _select_project_in_projects_tab(self, project_id):
        # go to Projects tab
        self.notebook.select(self.tab_tasks)

        if not hasattr(self, "task_tree"):
            return

        # search tree rows for matching ID
        for iid in self.task_tree.get_children():
            vals = self.task_tree.item(iid, "values")
            if not vals:
                continue
            if str(vals[0]) == str(project_id):
                self.task_tree.selection_set(iid)
                self.task_tree.focus(iid)
                self.task_tree.see(iid)
                return

    def _on_alert_double_click(self, event=None):
        sel = self.alert_tree.selection()
        if not sel:
            return
        vals = self.alert_tree.item(sel[0], "values")
        if not vals:
            return
        project_id = vals[1]
        self._select_project_in_projects_tab(project_id)

    def _on_near_double_click(self, event=None):
        sel = self.near_tree.selection()
        if not sel:
            return
        vals = self.near_tree.item(sel[0], "values")
        if not vals:
            return
        project_id = vals[0]
        self._select_project_in_projects_tab(project_id)

    def _add_dynamic_banner_grid(self, parent, image_path, row, col, height=28, padx=(8, 8)):
        lbl = ttk.Label(parent)
        lbl.grid(row=row, column=col, sticky="ew", padx=padx)
        parent.grid_columnconfigure(col, weight=1)

        if not os.path.exists(image_path):
            return lbl

        base_img = Image.open(image_path).convert("RGBA")
        lbl._banner_base = base_img
        lbl._banner_height = height
        lbl._banner_last_w = 0
        lbl._banner_job = None

        def redraw():
            lbl._banner_job = None
            w = max(10, lbl.winfo_width())
            if w == lbl._banner_last_w:
                return
            lbl._banner_last_w = w

            h = lbl._banner_height
            img = lbl._banner_base

            scale = max(w / img.width, h / img.height)
            new_w = int(img.width * scale)
            new_h = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            left = (new_w - w) // 2
            top  = (new_h - h) // 2
            cropped = resized.crop((left, top, left + w, top + h))

            photo = ImageTk.PhotoImage(cropped)
            lbl.configure(image=photo)
            lbl.image = photo

        def on_resize(event=None):
            if lbl._banner_job is not None:
                lbl.after_cancel(lbl._banner_job)
            lbl._banner_job = lbl.after(30, redraw)

        parent.bind("<Configure>", on_resize)
        lbl.after(50, redraw)
        return lbl
    
    def _add_dynamic_banner_pack(self, parent, image_path, height=28, padx=(8, 8)):
        lbl = ttk.Label(parent)
        lbl.pack(side="left", fill="x", expand=True, padx=padx)

        if not os.path.exists(image_path):
            return lbl  # acts as a spacer

        base_img = Image.open(image_path).convert("RGBA")
        lbl._banner_base = base_img
        lbl._banner_height = height
        lbl._banner_last_w = 0
        lbl._banner_job = None

        def redraw():
            lbl._banner_job = None
            w = max(10, lbl.winfo_width())
            if w == lbl._banner_last_w:
                return
            lbl._banner_last_w = w

            h = lbl._banner_height
            img = lbl._banner_base

            # Cover (w x h) then crop center (keeps horizontal look)
            scale = max(w / img.width, h / img.height)
            new_w = int(img.width * scale)
            new_h = int(img.height * scale)
            resized = img.resize((new_w, new_h), Image.LANCZOS)

            left = (new_w - w) // 2
            top  = (new_h - h) // 2
            cropped = resized.crop((left, top, left + w, top + h))

            photo = ImageTk.PhotoImage(cropped)
            lbl.configure(image=photo)
            lbl.image = photo  # keep reference

        def on_resize(event=None):
            if lbl._banner_job is not None:
                lbl.after_cancel(lbl._banner_job)
            lbl._banner_job = lbl.after(30, redraw)

        # bind to the label itself (avoids multiple parent bindings)
        lbl.bind("<Configure>", on_resize)
        lbl.after(50, redraw)
        return lbl

    def _get_active_table(self) -> str:
    # default dataset table
        return getattr(self, "current_tasks_table", "tasks")

    def load_selected_dataset(self, event=None):
        display = self.dataset_var.get()
        table = self.dataset_map.get(display, "tasks")

        self.current_tasks_table = table
        self.load_tasks()
        self.refresh_dashboard()

        # OPTIONAL: just clear CPM/Queueing views, don't compute
        self.cpm_tree.delete(*self.cpm_tree.get_children())
        self.queue_tree.delete(*self.queue_tree.get_children())

    # clear gantt widgets
        for w in self.cpm_chart_frame.winfo_children():
            w.destroy()
    
    def set_user(self, email, role):
        self.current_user = email
        self.current_role = role

        display_role = self.current_role.capitalize()

        last_refresh = datetime.now().strftime("%I:%M %p")

        self.user_label.config(
            text=(
                f"Logged in: {self.current_user}\n"
                f"Role: {display_role}\n"
                f"Last refresh: {last_refresh}"
            )
        )

    def refresh_user_session(self):
        display_role = (self.current_role or "").capitalize()
        last_refresh = datetime.now().strftime("%I:%M %p")

        self.user_label.config(
            text=(
                f"Logged in: {self.current_user}\n"
                f"Role: {display_role}\n"
                f"Last refresh: {last_refresh}"
            )
        )

    # ---------------- TASKS CRUD ----------------
    def load_tasks(self):
        for row in self.task_tree.get_children():
            self.task_tree.delete(row)

        table_name = self._get_active_table()

        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()

        q = sql.SQL("""
            SELECT
                ID, CUSTOMER, PROJECT_NAME, DATE_OF_PO, LEADTIME, STATUS,
                TARGET_DELIVERY, DEPENDENCIES, LAMBDA_VAL, MU_VAL, NOS_VAL,
                WORK_TYPE, REMARKS
            FROM {tbl}
            ORDER BY ID
        """).format(tbl=sql.Identifier(table_name))

        cur.execute(q)
        rows = cur.fetchall()

        cur.close()
        conn.close()

        for row in rows:
            status_val = row[5] if len(row) > 5 else None
            status_tag = str(status_val).strip().upper() if status_val else ""

            tags = ()
            if status_tag in STATUS_COLORS:
                tags = (status_tag,)

            self.task_tree.insert("", "end", values=row, tags=tags)

    def add_task(self):
        self.task_window()

    def edit_task(self):
        selected = self.task_tree.focus()
        if not selected:
            messagebox.showerror("Error", "Select a task to edit")
            return
        values = self.task_tree.item(selected, "values")
        self.task_window(values)

    def delete_task(self):
        selected = self.task_tree.focus()
        if not selected:
            messagebox.showerror("Error", "Select a task to delete")
            return

        task_id = int(self.task_tree.item(selected, "values")[0])
        table_name = self._get_active_table()

        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()

        try:
            # --- Check for dependent tasks in the SAME active table ---
            q_check = sql.SQL("""
                SELECT ID, PROJECT_NAME, DEPENDENCIES
                FROM {tbl}
                WHERE DEPENDENCIES IS NOT NULL
            """).format(tbl=sql.Identifier(table_name))

            cur.execute(q_check)
            rows = cur.fetchall()

            dependent_tasks = []
            for tid, tname, deps in rows:
                if not deps:
                    continue
                dep_ids = [int(x.strip()) for x in str(deps).split(",") if x.strip().isdigit()]
                if task_id in dep_ids:
                    dependent_tasks.append((tid, tname))

            if dependent_tasks:
                dep_list = "\n".join([f"ID {tid}: {tname}" for tid, tname in dependent_tasks])
                proceed = messagebox.askyesno(
                    "Dependent Tasks Found",
                    f"The following tasks depend on the task you are trying to delete (ID {task_id}):\n\n"
                    f"{dep_list}\n\n"
                    "Deleting this task will leave these dependencies broken.\n"
                    "Do you want to proceed anyway?"
                )
                if not proceed:
                    return

            # --- Delete from SAME active table ---
            q_del = sql.SQL("DELETE FROM {tbl} WHERE id=%s").format(tbl=sql.Identifier(table_name))
            cur.execute(q_del, (task_id,))
            conn.commit()

        except Exception as e:
            conn.rollback()
            messagebox.showerror("Delete Error", str(e))
            return

        finally:
            cur.close()
            conn.close()

        self.load_tasks()
  
    # ---------------- TASK WINDOW ----------------
    def task_window(self, values=None):
        win = tk.Toplevel(self)
        win.title("PROJECT")
        win.geometry("350x800")

    # --- ID Field ---
        ttk.Label(win, text="TASK ID:").pack(pady=5)
        id_entry = ttk.Entry(win)
        id_entry.pack()

    # --- Task ID & Customer & Date of PO & Name & LEADTIME & Status & Target Delivery & Dependencies & lambda & mu & no. of server & remarks ---
        ttk.Label(win, text="CUSTOMER:").pack(pady=5)
        cust_name_entry = ttk.Entry(win)
        cust_name_entry.pack()
        
        ttk.Label(win, text="PROJECT NAME:").pack(pady=5)
        proj_name_entry = ttk.Entry(win)
        proj_name_entry.pack()

        ttk.Label(win, text="DATE OF PO:").pack(pady=5)
        date_po_entry = DateEntry(
            win,
            date_pattern="yyyy-mm-dd",   # matches your parser / PostgreSQL DATE
            width=18,
            state="readonly"
        )
        date_po_entry.pack()

        ttk.Label(win, text="LEADTIME: (Span of Weeks)").pack(pady=5)
        dur_entry = ttk.Entry(win)
        dur_entry.pack()

        ttk.Label(win, text="STATUS:").pack(pady=5)
        STATUS_OPTIONS = [
            "DELAY",
            "ON-HOLD",
            "DELIVERED",
            "COMPLETED",
            "ONGOING",
            "SUBJECT FOR MODIFICATION",
            "ON TIME"
        ]
        
        status_var = tk.StringVar()
        status_combo = ttk.Combobox(
            win,
            textvariable=status_var,
            values=STATUS_OPTIONS,
            state="readonly",
            width=18
        )
        status_combo.pack(pady=2)

        ttk.Label(win, text="TARGET DELIVERY:").pack(pady=5)
        del_entry = DateEntry(
            win,
            date_pattern="yyyy-mm-dd",
            width=18,
            state="readonly"
        )
        del_entry.pack()

        ttk.Label(win, text="DEPENDENCIES (comma IDs):").pack(pady=5)
        dep_entry = ttk.Entry(win)
        dep_entry.pack()

        ttk.Label(win, text="Lambda:").pack(pady=5)
        lam_entry = ttk.Entry(win)
        lam_entry.pack()

        ttk.Label(win, text="Mu:").pack(pady=5)
        mu_entry = ttk.Entry(win)
        mu_entry.pack()

        ttk.Label(win, text="No. of Server: (c)").pack(pady=5)
        no_of_server_entry = ttk.Entry(win)
        no_of_server_entry.pack()

        ttk.Label(win, text="WORK TYPE:").pack(pady=5)
        WORK_OPTIONS = [
            "PURCHASE ORDER",
            "PROJECT",
            "CHANGE REQUEST",
            "SERVICE JOB",
            "MAINTENANCE",
            "CUSTOM"
        ]
        
        work_var = tk.StringVar()
        work_combo = ttk.Combobox(
            win,
            textvariable=work_var,
            values=WORK_OPTIONS,
            state="readonly",
            width=18
        )
        work_combo.pack(pady=2)

        ttk.Label(win, text="REMARKS: ").pack(pady=5)
        remarks_entry = ttk.Entry(win)
        remarks_entry.pack()

    # --- Prefill values if editing ---
        if values:
            id_entry.insert(0, values[0])
            id_entry.config(state='disabled')  # ID should not be changed when editing
            cust_name_entry.insert(0, values[1])
            proj_name_entry.insert(0, values[2])
            if values and values[3]:
                date_po_entry.set_date(values[3])
            dur_entry.insert(0, values[4])
            status_var.set(values[5])
            if values and values[6]:
                del_entry.set_date(values[6])
            dep_entry.insert(0, values[7])
            lam_entry.insert(0, values[8])
            mu_entry.insert(0, values[9])
            no_of_server_entry.insert(0, values[10])
            work_var.set(values[11])
            remarks_entry.insert(0, values[12])
            
        else:
            # Optional: auto-fill ID for new task
            conn = psycopg2.connect(**DB_PARAMS)
            cur = conn.cursor()
            table_name = self._get_active_table()
            cur.execute(
                sql.SQL("SELECT COALESCE(MAX(ID), 0) FROM {tbl}")
                .format(tbl=sql.Identifier(table_name))
            )
            max_id = cur.fetchone()[0] or 0
            cur.close()
            conn.close()
            id_entry.insert(0, str(max_id + 1))

        def save_task():
            try:
                task_id = int(id_entry.get().strip())
            except ValueError:
                messagebox.showwarning("Invalid ID", "Task ID must be a numeric value.")
                return

            cust = cust_name_entry.get().strip()
            name = proj_name_entry.get().strip()
            date = date_po_entry.get()
            dur_raw = dur_entry.get().strip()
            status = status_var.get()
            deliver = del_entry.get()
            dep = dep_entry.get().strip()
            lam = lam_entry.get().strip()
            mu = mu_entry.get().strip()
            nos = no_of_server_entry.get().strip()
            work = work_var.get()
            remarks =  remarks_entry.get().strip()

        # Required fields
            if not cust or not name or not dur_raw or not date:
                messagebox.showwarning(
                    "Cannot Save Task",
                    "Customer, Task Name, Leadtime, and Date of PO are required."
                )
                return
            
            try:
                po_date_val = self.parse_date_or_none(date)
                if po_date_val is None:
                    raise ValueError("Date of PO is required.")
            except ValueError as e:
                messagebox.showwarning("Invalid Date of PO", str(e))
                return

        # Target delivery is OPTIONAL
            try:
                delivery_val = self.parse_date_or_none(deliver)  # may become None
            except ValueError as e:
                messagebox.showwarning("Invalid Target Delivery", str(e))
                return

        # LEADTIME validation
            try:
                dur_days = self.weeks_span_to_days(dur_raw)
            except ValueError as e:
                messagebox.showwarning("Invalid LEADTIME", str(e))
                return

        # Lambda & Mu validation & Number of Server
            try:
                lam_val = float(lam) if lam else None
                mu_val = float(mu) if mu else None
                nos_val = int(nos) if nos else None
            except ValueError:
                messagebox.showwarning(
                    "Invalid Input",
                    "Lambda, Mu, and Number of Server must be numeric values."
                )
                return

            work = work_var.get().strip()
            if not work:
                messagebox.showwarning("Missing Work Type", "Please select a WORK TYPE.")
                return

            status = status_var.get().strip()
            if not status:
                messagebox.showwarning("Missing Status", "Please select a STATUS.")
                return
            
                # Remarks OPTIONAL
            remarks_val = remarks.strip() if remarks.strip() else None

        # Save to DB (Option B: save to active dataset table)
            table_name = self._get_active_table()

            conn = psycopg2.connect(**DB_PARAMS)
            cur = conn.cursor()
            try:
                if values:  # Editing existing task
                    q = sql.SQL("""
                        UPDATE {tbl}
                        SET CUSTOMER=%s,
                            PROJECT_NAME=%s,
                            DATE_OF_PO=%s,
                            LEADTIME=%s,
                            STATUS=%s,
                            TARGET_DELIVERY=%s,
                            DEPENDENCIES=%s,
                            LAMBDA_VAL=%s,
                            MU_VAL=%s,
                            NOS_VAL=%s,
                            WORK_TYPE=%s,
                            REMARKS=%s
                        WHERE ID=%s
                    """).format(tbl=sql.Identifier(table_name))

                    cur.execute(q, (
                        cust, name, po_date_val, dur_days, status, delivery_val,
                        dep, lam_val, mu_val, nos_val, work, remarks_val, task_id
                    ))

                else:  # New task
                    q = sql.SQL("""
                        INSERT INTO {tbl} (
                            ID, CUSTOMER, PROJECT_NAME, DATE_OF_PO, LEADTIME, STATUS,
                            TARGET_DELIVERY, DEPENDENCIES,
                            LAMBDA_VAL, MU_VAL, NOS_VAL,
                            WORK_TYPE, REMARKS
                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """).format(tbl=sql.Identifier(table_name))

                    cur.execute(q, (
                        task_id, cust, name, po_date_val, dur_days, status, delivery_val,
                        dep, lam_val, mu_val, nos_val, work, remarks_val
                    ))

                conn.commit()

            except Exception as e:
                conn.rollback()
                messagebox.showerror("Database Error", str(e))
                return

            finally:
                cur.close()
                conn.close()

            self.load_tasks()
            win.destroy()
        ttk.Button(win, text="Save", command=save_task).pack(pady=15)

    # ---------------- LEADTIME PARSER ----------------
    def weeks_span_to_days(self, duration_str):
        clean = duration_str.lower().replace("weeks", "").replace("week", "").strip()

        if not clean:
            raise ValueError("LEADTIME cannot be empty.")

        try:
            if "-" in clean:
                min_w, max_w = clean.split("-")
                avg_weeks = (float(min_w.strip()) + float(max_w.strip())) / 2
            else:
                avg_weeks = float(clean)

            return avg_weeks * 7
        except ValueError:
            raise ValueError("Invalid LEADTIME format. Use e.g. '4-6 weeks'.")
    
    def edit_task_by_id(self, task_id):
        """
        Open the edit window for a specific task ID (no Treeview selection needed),
        using the currently selected dataset table.
        """
        table_name = self._get_active_table()

        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()

        q = sql.SQL("""
            SELECT ID, CUSTOMER, PROJECT_NAME, LEADTIME, DEPENDENCIES
            FROM {tbl}
            WHERE ID = %s
        """).format(tbl=sql.Identifier(table_name))

        cur.execute(q, (task_id,))
        row = cur.fetchone()

        cur.close()
        conn.close()

        if not row:
            messagebox.showerror("Error", f"Task {task_id} not found in {table_name}.")
            return

        values = (
            row[0],  # ID
            row[1],  # CUSTOMER
            row[2],  # PROJECT_NAME
            row[3],  # LEADTIME
            row[4],  # DEPENDENCIES
        )

        self.task_window(values)
    
    def parse_date_or_none(self, date_str):
        date_str = date_str.strip()
        if not date_str:
            return None
    # Expect YYYY-MM-DD (best with PostgreSQL DATE)
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            raise ValueError("Use date format YYYY-MM-DD (example: 2025-07-22)")
        
    def _fix_dateentry_text(self, event):
        w = event.widget
        try:
            d = w.get_date()  # datetime.date
            w.delete(0, "end")
            w.insert(0, d.strftime("%Y-%m-%d"))
        except Exception:
            pass

    def _delete_task_by_id(self, task_id):
        table_name = self._get_active_table()

        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()
        try:
            q = sql.SQL("DELETE FROM {tbl} WHERE id=%s").format(
                tbl=sql.Identifier(table_name)
            )
            cur.execute(q, (task_id,))
            conn.commit()
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Delete Error", str(e))
            return
        finally:
            cur.close()
            conn.close()

        messagebox.showinfo("Deleted", f"Task {task_id} removed from {table_name}.")
        self.load_tasks()

    def _missing_deps_dialog(self, missing):
        """
        missing: list of tuples (tid, name, dep)
        returns: (action, chosen_tid)
        action in {"edit","remove","ignore"}
        chosen_tid is the selected task id (or None)
        """

        # Use your app root safely even if you don't store self.root
        parent = getattr(self, "root", None)
        if parent is None:
            parent = self.cpm_chart_frame.winfo_toplevel()

        dialog = tk.Toplevel(parent)
        dialog.title("Missing Dependencies Found")
        dialog.transient(parent)
        dialog.grab_set()
        dialog.resizable(False, False)

        result = {"action": "ignore", "tid": None}

        ttk.Label(
            dialog,
            text="Some tasks have missing dependencies.\nSelect a task, then choose an action:",
            font=("Segoe UI", 10, "bold")
        ).pack(padx=12, pady=(12, 6), anchor="w")

        # Build unique task list from 'missing'
        # Example: Task 5 might have 2 missing deps; we show it once.
        task_to_lines = {}
        for tid, name, dep in missing:
            task_to_lines.setdefault(tid, {"name": name, "deps": []})
            task_to_lines[tid]["deps"].append(dep)

        items = sorted(task_to_lines.items(), key=lambda kv: kv[0])  # sort by tid

        lb = tk.Listbox(dialog, width=72, height=min(10, max(5, len(items))))
        lb.pack(padx=12, pady=6)

        # store tids in same order as listbox
        tids = []
        for tid, info in items:
            deps_str = ", ".join(str(d) for d in sorted(set(info["deps"])))
            lb.insert("end", f"Task {tid} ({info['name']}) → Missing: {deps_str}")
            tids.append(tid)

        if tids:
            lb.selection_set(0)

        hint = ttk.Label(dialog, text="Tip: 'Ignore' will remove only the invalid dependency links and continue CPM.")
        hint.pack(padx=12, pady=(0, 8), anchor="w")

        btns = ttk.Frame(dialog)
        btns.pack(padx=12, pady=(0, 12), anchor="e")

        def get_selected_tid():
            sel = lb.curselection()
            if not sel:
                return None
            return tids[sel[0]]

        def do_edit():
            tid = get_selected_tid()
            if tid is None:
                messagebox.showwarning("Select a Task", "Please select a task first.", parent=dialog)
                return
            result["action"] = "edit"
            result["tid"] = tid
            dialog.destroy()

        def do_remove():
            tid = get_selected_tid()
            if tid is None:
                messagebox.showwarning("Select a Task", "Please select a task first.", parent=dialog)
                return
            if messagebox.askyesno(
                "Confirm Delete",
                f"Delete Task {tid}? This cannot be undone.",
                parent=dialog
            ):
                result["action"] = "remove"
                result["tid"] = tid
                dialog.destroy()

        def do_ignore():
            result["action"] = "ignore"
            result["tid"] = None
            dialog.destroy()

        ttk.Button(btns, text="Edit Task", command=do_edit).pack(side="left", padx=6)
        ttk.Button(btns, text="Remove Task", command=do_remove).pack(side="left", padx=6)
        ttk.Button(btns, text="Ignore", command=do_ignore).pack(side="left", padx=6)

        dialog.wait_window()
        return result["action"], result["tid"]

    # ---------------- CPM ----------------
    def calculate_cpm(self):
    # --- 1. Fetch tasks ---
        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()
        tbl = self._get_active_table()
        q = sql.SQL("""
            SELECT ID, CUSTOMER, PROJECT_NAME, LEADTIME, DEPENDENCIES
            FROM {tbl}
            ORDER BY ID
        """).format(tbl=sql.Identifier(tbl))
        cur.execute(q)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if not rows:
            messagebox.showwarning("No Tasks", "There are no tasks to calculate CPM.")
            return None

        # --- 2. Build task dictionary ---
        tasks = {}
        for r in rows:
            tid = int(r[0])
            customer = r[1]
            name = r[2]
            DURATION = float(r[3])

            dep_str = r[4] or ""
            deps = []
            if dep_str.strip():
                deps = [int(x.strip()) for x in dep_str.split(",") if x.strip()]

            tasks[tid] = {
                "PROJECT NAME": name,
                "CUSTOMER": customer,
                "DURATION": DURATION,
                "DEPS": deps,
                "SUCCS": [],
                "ES": 0.0,
                "EF": 0.0,
                "LS": 0.0,
                "LF": 0.0,
                "SLACK": 0.0,
                "CRITICAL": False
            }

        # --- 3. Build successors & check missing dependencies ---
        # --- 3. Build successors & handle missing dependencies ---
        missing = []
        for tid, t in tasks.items():
            cleaned = []
            for dep in t["DEPS"]:
                if dep in tasks:
                    tasks[dep]["SUCCS"].append(tid)
                    cleaned.append(dep)
                else:
                    missing.append((tid, t["PROJECT NAME"], dep))
            t["DEPS"] = cleaned  # keep only valid deps

        if missing:
            action, chosen_tid = self._missing_deps_dialog(missing)

            if action == "edit" and chosen_tid is not None:
                self.edit_task_by_id(chosen_tid)   # <-- hook to your edit UI
                return None  # stop CPM until they fix

            if action == "remove" and chosen_tid is not None:
                self._delete_task_by_id(chosen_tid)
                return None  # stop CPM; user can rerun CPM after

    # action == "ignore"
    # continue CPM, but deps are already cleaned below

        # --- 4. Topological sort ---
        in_deg = {tid: len(t["DEPS"]) for tid, t in tasks.items()}
        queue = deque([tid for tid, d in in_deg.items() if d == 0])
        topo = []

        while queue:
            u = queue.popleft()
            topo.append(u)
            for s in tasks[u]["SUCCS"]:
                in_deg[s] -= 1
                if in_deg[s] == 0:
                    queue.append(s)
                
        if len(topo) != len(tasks):
            missing_ids = sorted(set(tasks.keys()) - set(topo))
            messagebox.showwarning("Circular Dependency",
                f"These tasks are in a dependency loop (not schedulable): {missing_ids}"
            )
            
            bad = set(tasks.keys()) - set(topo)
            for b in bad:
                # remove them from graph so they don't break plotting
                tasks.pop(b, None)

            return tasks

        # Cycle detection
        if len(topo) != len(tasks):
            messagebox.showwarning(
                "Circular Dependency",
                "One or more tasks are in a dependency loop.\n\nFix before running CPM."
            )
            return None

        # --- 5. Forward pass ---
        for tid in topo:
            t = tasks[tid]
            t["ES"] = 0 if not t["DEPS"] else max(tasks[d]["EF"] for d in t["DEPS"])
            t["EF"] = t["ES"] + t["DURATION"]

        # --- 6. Backward pass ---
        max_EF = max(t["EF"] for t in tasks.values())

        for tid in reversed(topo):
            t = tasks[tid]
            t["LF"] = max_EF if not t["SUCCS"] else min(tasks[s]["LS"] for s in t["SUCCS"])
            t["LS"] = t["LF"] - t["DURATION"]
            t["SLACK"] = t["LS"] - t["ES"]
            t["CRITICAL"] = abs(t["SLACK"]) < 1e-9

        return tasks

# ---------------- Display CPM in Tab ----------------
    def load_all_tasks_in_cpm_tab(self):
    # clear existing rows
        self.cpm_all_tasks_tree.delete(*self.cpm_all_tasks_tree.get_children())

        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()
        tbl = self._get_active_table()
        q = sql.SQL("""
            SELECT
                ID,
                CUSTOMER,
                PROJECT_NAME,
                DATE_OF_PO,
                LEADTIME,
                STATUS,
                TARGET_DELIVERY,
                DEPENDENCIES,
                LAMBDA_VAL,
                MU_VAL,
                NOS_VAL,
                WORK_TYPE,
                REMARKS
            FROM {tbl}
            ORDER BY ID
        """).format(tbl=sql.Identifier(tbl))
        cur.execute(q)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        for r in rows:
            self.cpm_all_tasks_tree.insert("", "end", values=r)

    def refresh_key_metrics(self):
        table = getattr(self, "current_tasks_table", "tasks")

        sql = f"""
            SELECT
                AVG(leadtime) FILTER (WHERE leadtime IS NOT NULL),
                ROUND(
                    100.0 * SUM(CASE WHEN UPPER(status) = 'ON TIME' THEN 1 ELSE 0 END)
                    / NULLIF(COUNT(*), 0), 1
                ),
                AVG(leadtime) FILTER (WHERE UPPER(status) = 'DELAY'),
                COUNT(*) FILTER (WHERE UPPER(status) = 'DELAY')
            FROM {table}
        """

        with self._db() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                avg_lead, ontime_rate, avg_delay, critical = cur.fetchone()

        self.lbl_avg_lead.config(
            text=f"Avg Lead Time: {avg_lead:.1f} days" if avg_lead else "Avg Lead Time: --"
        )
        self.lbl_on_time.config(
            text=f"On-Time Rate: {ontime_rate}%" if ontime_rate is not None else "On-Time Rate: --%"
        )
        self.lbl_avg_delay.config(
            text=f"Avg Delay: {avg_delay:.1f} days" if avg_delay else "Avg Delay: --"
        )
        self.lbl_critical.config(
            text=f"Critical Projects: {critical}"
        )

    def show_cpm_table_in_tab(self):
        self.load_all_tasks_in_cpm_tab()
        self.cpm_tree.delete(*self.cpm_tree.get_children())

        tasks = self.calculate_cpm()
        if tasks is None:
            return

        for tid, t in tasks.items():
            tag = "critical" if t["CRITICAL"] else "normal"

            deps_display = ",".join(str(d) for d in t["DEPS"]) if t["DEPS"] else ""

            self.cpm_tree.insert(
                "",
                "end",
                values=(
                    tid,
                    t["PROJECT NAME"],
                    t["CUSTOMER"],
                    t["DURATION"],
                    deps_display,
                    t["ES"],
                    t["EF"],
                    t["LS"],
                    t["LF"],
                    t["SLACK"],
                    "YES" if t["CRITICAL"] else "NO"
                ),
                tags=(tag,)
            )  
    
    def draw_cpm_gantt(self):
        # ALWAYS clear chart frame first so you never see stale charts
        for w in self.cpm_chart_frame.winfo_children():
            w.destroy()

        tasks = self.calculate_cpm()
        if not tasks:
            ttk.Label(
                self.cpm_chart_frame,
                text="Cannot draw Gantt: CPM returned no data (check missing/circular dependencies)."
            ).pack(padx=10, pady=10, anchor="w")
            return

        # sort by ES then ID
        items = sorted(tasks.items(), key=lambda kv: (kv[1]["ES"], kv[0]))

        y_labels, es, dur, crit = [], [], [], []
        for tid, t in items:
            y_labels.append(f"{tid} | {t['CUSTOMER']} | {t['PROJECT NAME']}")
            es.append(float(t["ES"]))
            dur.append(float(t["DURATION"]))
            crit.append(bool(t["CRITICAL"]))

        fig = plt.Figure(figsize=(10, 5), dpi=100)
        ax = fig.add_subplot(111)

        y = list(range(len(items)))
        ax.barh(y, dur, left=es, height=0.6)

        # hatch critical tasks
        for i, is_crit in enumerate(crit):
            if is_crit:
                ax.barh(i, dur[i], left=es[i], height=0.6, hatch="///", fill=False)

        ax.set_yticks(y)
        ax.set_yticklabels(y_labels, fontsize=8)
        ax.set_xlabel("Time")
        ax.set_title("CPM Gantt Chart (ES→EF), critical tasks hatched")
        ax.invert_yaxis()

        project_finish = max(t["EF"] for _, t in items)
        ax.axvline(project_finish, linestyle="--")
        ax.text(project_finish, -0.5, f"Finish = {project_finish}", fontsize=8)

        # ✅ ADD THIS LINE so export can grab the figure
        self.cpm_gantt_fig = fig

        self.cpm_canvas = FigureCanvasTkAgg(fig, master=self.cpm_chart_frame)
        self.cpm_canvas.draw()
        self.cpm_canvas.get_tk_widget().pack(fill="both", expand=True)
        
    # ---------------- QUEUEING ----------------
    def _fetch_queue_tasks(self):
        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()
        tbl = self._get_active_table()
        q = sql.SQL("""
            SELECT ID, CUSTOMER, PROJECT_NAME, WORK_TYPE, LAMBDA_VAL, MU_VAL, NOS_VAL
            FROM {tbl}
            ORDER BY ID
        """).format(tbl=sql.Identifier(tbl))
        cur.execute(q)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return rows


    def _mmc_metrics(self, lam, mu, c):
        """
        Simple M/M/c metrics.
        Returns tuple:
        (rho, wq, w, lq, l, p0, note)
        """
        note = ""

        if lam is None or mu is None or c is None:
            return None, None, None, None, None, None, "Missing λ/μ/c"

        try:
            lam = float(lam)
            mu = float(mu)
            c = int(c)
        except Exception:
            return None, None, None, None, None, None, "Invalid λ/μ/c"

        if mu <= 0 or c <= 0:
            return None, None, None, None, None, None, "μ and c must be > 0"

        rho = lam / (c * mu)

        # if unstable, return rho (so we can tag it) + blanks
        if rho >= 1:
            return rho, None, None, None, None, None, "UNSTABLE (ρ ≥ 1)"

        # Erlang C
        a = lam / mu
        s = 0.0
        for n in range(c):
            s += (a ** n) / math.factorial(n)
        last = (a ** c) / (math.factorial(c) * (1 - rho))
        p0 = 1.0 / (s + last)

        pw = last * p0
        lq = (pw * rho) / (1 - rho)
        wq = lq / lam if lam > 0 else 0.0
        w = wq + (1 / mu)
        l = lam * w

        return rho, wq, w, lq, l, p0, note


    def _clear_queue_tree(self):
        for row in self.queue_tree.get_children():
            self.queue_tree.delete(row)


    def _insert_queue_row(self, tid, customer, name, work_type, lam, mu, c, rho, wq, w, lq, l, p0, note, tag=None):
        def fmt(x):
            if x is None:
                return ""
            if isinstance(x, float):
                return f"{x:.4f}"
            return str(x)

        # IMPORTANT: order must match queue_cols
        self.queue_tree.insert(
            "", "end",
            values=(
                tid,
                customer,
                name,           # PROJECT NAME
                work_type,      # WORK TYPE
                fmt(lam),
                fmt(mu),
                fmt(c),
                fmt(rho),
                fmt(wq),
                fmt(w),
                fmt(lq),
                fmt(l),
                fmt(p0),
                note or ""
            ),
            tags=(tag,) if tag else ()
        )

    def show_queue_all(self):
        try:
            self._render_queue_table(grouped=False)
        except Exception:
            traceback.print_exc()
            messagebox.showerror("Queue Error", traceback.format_exc())

    def show_queue_grouped(self):
        self._render_queue_table(grouped=True)


    def _render_queue_table(self, grouped: bool):
        # tags (safe to call multiple times)
        self.queue_tree.tag_configure("group", background="#e6e6e6")
        self.queue_tree.tag_configure("unstable", background="#ffb3b3")
        self.queue_tree.tag_configure("warning", background="#fff2b3")

        rows = self._fetch_queue_tasks()
        if not rows:
            # optional: show a single row message
            self.queue_tree.insert("", "end", values=("", "", "No tasks found", "", "", "", "", "", "", "", "", "", "", ""))
            return

        computed = []
        for (tid, customer, name, work_type, lam, mu, c) in rows:
            wt = (work_type or "UNSPECIFIED").strip().upper()

            rho, wq, w, lq, l, p0, note = self._mmc_metrics(lam, mu, c)

            computed.append({
                "ID": tid,
                "CUSTOMER": customer,
                "PROJECT NAME": name,
                "WORK TYPE": wt,
                "LAMBDA": lam,
                "MU": mu,
                "SERVERS": c,
                "RHO": rho, "WQ": wq, "W": w, "LQ": lq, "L": l, "P0": p0,
                "NOTE": note
            })

        if not grouped:
            for r in computed:
                tag = None
                if r["RHO"] is not None and r["RHO"] >= 1:
                    tag = "unstable"
                elif r["RHO"] is not None and r["RHO"] >= 0.85:
                    tag = "warning"

                self._insert_queue_row(
                    r["ID"], r["CUSTOMER"], r["PROJECT NAME"], r["WORK TYPE"], r["LAMBDA"], r["MU"], r["SERVERS"],
                    r["RHO"], r["WQ"], r["W"], r["LQ"], r["L"], r["P0"], r["NOTE"],
                    tag=tag
                )
            return

        # GROUPED view
        computed.sort(key=lambda x: (x["WORK TYPE"], x["ID"]))
        current_group = None

        for r in computed:
            wt = r["WORK TYPE"] or "UNSPECIFIED"
            if wt != current_group:
                current_group = wt
                # group header row
                self.queue_tree.insert(
                    "", "end",
                    values=("", f"WORK TYPE: {wt}", "", "", "", "", "", "", "", "", "", "", ""),
                    tags=("group",)
                )

            tag = None
            if r["RHO"] is not None and r["RHO"] >= 1:
                tag = "unstable"
            elif r["RHO"] is not None and r["RHO"] >= 0.85:
                tag = "warning"

            self._insert_queue_row(
                r["ID"], r["CUSTOMER"], r["PROJECT NAME"], r["WORK TYPE"], r["LAMBDA"], r["MU"], r["SERVERS"],
                r["RHO"], r["WQ"], r["W"], r["LQ"], r["L"], r["P0"], r["NOTE"],
                tag=tag
            )
    
    def draw_queue_heatmap(self, summary_rows):
        # ALWAYS clear heatmap frame first (same idea as gantt)
        for w in self.heatmap_frame.winfo_children():
            w.destroy()

        # Expand heatmap row now that we're drawing
        self.tab_queueing.grid_rowconfigure(3, weight=1, minsize=220)

        # Create figure fresh each time (simple + consistent)
        fig = Figure(figsize=(10, 2.8), dpi=100)
        ax = fig.add_subplot(111)

        if not summary_rows:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", fontsize=16)
            ax.set_axis_off()
        else:
            work_types = [r["WORK_TYPE"] for r in summary_rows]
            data = [[r["AVG_RHO"], r["MAX_RHO"]] for r in summary_rows]

            im = ax.imshow(data, aspect="auto", vmin=0, vmax=1)

            ax.set_yticks(range(len(work_types)))
            ax.set_yticklabels(work_types)
            ax.set_xticks([0, 1])
            ax.set_xticklabels(["AVG_RHO", "MAX_RHO"])
            ax.set_title("Queue Utilization Heatmap")

            fig.subplots_adjust(left=0.2, right=0.92, bottom=0.25, top=0.85)
            cbar = fig.colorbar(im, ax=ax, fraction=0.04, pad=0.03)
            cbar.set_label("Utilization (ρ)")

        # ✅ save for excel export (like gantt)
        self.queue_heatmap_fig = fig

        self.queue_heatmap_canvas = FigureCanvasTkAgg(fig, master=self.heatmap_frame)
        self.queue_heatmap_canvas.draw()
        self.queue_heatmap_canvas.get_tk_widget().pack(fill="both", expand=True)

    def calculate_queueing(self):
        # HARD RESET: clear everything before inserting again
        self.queue_tree.delete(*self.queue_tree.get_children())
        # render detail table first (same logic as SHOW ALL)
        self._render_queue_table(grouped=False)

        # clear summary
        for row in self.queue_summary_tree.get_children():
            self.queue_summary_tree.delete(row)

        rows = self._fetch_queue_tasks()
        if not rows:
            self._last_summary_rows = []
            return

        agg = defaultdict(lambda: {
            "COUNT": 0,
            "TOTAL_LAMBDA": 0.0,
            "MU_LIST": [],
            "TOTAL_SERVERS": 0,
            "RHO_LIST": [],
            "MAX_RHO": 0.0,
            "FLAGS": []
        })

        for (tid, customer, name, work_type, lam, mu, servers) in rows:
            wt = (work_type or "UNSPECIFIED").strip().upper()

            rho, wq, w, lq, l, p0, note = self._mmc_metrics(lam, mu, servers)

            lam_f = float(lam) if lam is not None else 0.0
            mu_f = float(mu) if mu is not None else 0.0
            c_i = int(servers) if servers is not None else 1

            a = agg[wt]
            a["COUNT"] += 1
            a["TOTAL_LAMBDA"] += lam_f
            a["MU_LIST"].append(mu_f)
            a["TOTAL_SERVERS"] += c_i

            if rho is not None and rho < 1:
                a["RHO_LIST"].append(rho)
                a["MAX_RHO"] = max(a["MAX_RHO"], rho)
                if rho >= 0.85:
                    a["FLAGS"].append("HIGH UTILIZATION")
            else:
                a["FLAGS"].append("UNSTABLE/INVALID INPUT")

        summary_rows = []
        for wt, a in agg.items():
            avg_mu = (sum(a["MU_LIST"]) / len(a["MU_LIST"])) if a["MU_LIST"] else 0.0
            avg_rho = (sum(a["RHO_LIST"]) / len(a["RHO_LIST"])) if a["RHO_LIST"] else 0.0
            max_rho = a["MAX_RHO"]
            flag_text = ", ".join(sorted(set(a["FLAGS"]))) if a["FLAGS"] else ""

            if "UNSTABLE/INVALID INPUT" in flag_text:
                tag = "unstable"
            elif max_rho >= 0.85:
                tag = "warning"
            else:
                tag = "ok"

            self.queue_summary_tree.insert(
                "", "end",
                values=(
                    wt,
                    a["COUNT"],
                    f"{a['TOTAL_LAMBDA']:.4f}",
                    f"{avg_mu:.4f}",
                    a["TOTAL_SERVERS"],
                    f"{avg_rho:.4f}",
                    f"{max_rho:.4f}",
                    flag_text
                ),
                tags=(tag,)
            )

            summary_rows.append({"WORK_TYPE": wt, "AVG_RHO": avg_rho, "MAX_RHO": max_rho})

        summary_rows.sort(key=lambda r: r["WORK_TYPE"])
        self._last_summary_rows = summary_rows
        
    def show_heatmap(self):
        # toggle: if currently expanded -> collapse
        current_min = self.tab_queueing.grid_rowconfigure(3)["minsize"]
        if current_min and int(current_min) > 0:
            self.tab_queueing.grid_rowconfigure(3, weight=0, minsize=0)
            self.heatmap_frame.configure(height=35)
            return

        # else expand and draw
        self.tab_queueing.grid_rowconfigure(3, weight=1, minsize=220)
        self.heatmap_frame.configure(height=260)

        rows = getattr(self, "_last_summary_rows", [])
        self.draw_queue_heatmap(rows)

    def run_delay_alerts(self):
        """
        Sends email alerts for delayed tasks.
        - Sends to Admin emails
        - Writes the Admin email (username) inside the email body
        - Anti-spam: once per day per task
        """
        table = getattr(self, "current_tasks_table", "tasks")

        try:
            delayed_rows = fetch_delayed_tasks(table)
            if not delayed_rows:
                return

            admin_emails = fetch_admin_emails()
            if not admin_emails:
                return

            for task_id, project_name, status, target_del in delayed_rows:
                # anti-spam check
                if not should_send_alert_today(task_id, "DELAYED"):
                    continue

                subject = f"[ALERT] Delayed Project: {project_name}"

                # send ONE email per admin (clear + traceable)
                for admin_email in admin_emails:
                    body = (
                        "Project Management System - Delay Alert\n\n"
                        f"Administrator: {admin_email}\n\n"
                        f"Project ID: {task_id}\n"
                        f"Project Name: {project_name}\n"
                        f"Status: {status}\n"
                        f"Target Delivery: {target_del}\n\n"
                        "Reason: Target delivery date has passed and the project "
                        "is not marked as Completed.\n"
                    )

                    send_email(admin_email, subject, body)

                mark_alert_sent(task_id, "DELAYED")

        except Exception as e:
            print("Email alert error:", e)


    def start_alert_loop(self):
        """
        Runs delay alert checks every 60 minutes.
        Call this ONCE after dashboard loads / after login.
        """
        self.run_delay_alerts()
        self.after(60 * 60 * 1000, self.start_alert_loop)  # 60 minutes

    # ---------------- EXCEL ----------------
    def _save_fig_to_png(self, fig, filename):
        fig.savefig(filename, dpi=150, bbox_inches="tight")

    def _sanitize_table_name(self, name: str) -> str:
        name = (name or "").strip()
        name = re.sub(r"[^a-zA-Z0-9_]+", "_", name)   # keep only safe chars
        name = name.lower()                          # IMPORTANT: lowercase always
        name = re.sub(r"_+", "_", name).strip("_")   # cleanup
        if not name:
            name = "dataset"
        if name[0].isdigit():
            name = "d_" + name
        return name
    
    def _unique_table_name(self, cur, base_name: str) -> str:
        base_name = self._sanitize_table_name(base_name)  # ensure lowercase
        candidate = base_name
        i = 1

        while True:
            cur.execute("""
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = 'public'
                AND table_name = %s
                LIMIT 1
            """, (candidate,))
            exists = cur.fetchone() is not None

            if not exists:
                return candidate

            i += 1
            candidate = f"{base_name}_{i}"

    def import_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path:
            return

        # 1) read raw to find header row
        try:
            raw = pd.read_excel(path, header=None)
        except Exception as e:
            messagebox.showerror("Import Error", f"Cannot read Excel file:\n{e}")
            return

        header_row = None
        for i in range(min(30, len(raw))):
            row_vals = (
                raw.iloc[i].astype(str)
                .str.replace("\n", " ", regex=False)
                .str.strip()
                .tolist()
            )
            # must match your exported headers
            if ("ID" in row_vals) and ("CUSTOMER" in row_vals) and ("PROJECT NAME" in row_vals):
                header_row = i
                break

        if header_row is None:
            messagebox.showerror(
                "Import Error",
                "Could not find the header row (ID, CUSTOMER, PROJECT NAME)."
            )
            return

        # 2) read with correct header row
        df = pd.read_excel(path, header=header_row)
        df.columns = (
            df.columns.astype(str)
            .str.replace("\n", " ", regex=False)
            .str.strip()
        )

        # 3) validate required columns
        required = {
            "ID", "CUSTOMER", "PROJECT NAME", "DATE OF PO", "LEADTIME",
            "LAMBDA", "MU", "NO. OF SERVER", "WORK TYPE"
        }
        missing = required - set(df.columns)
        if missing:
            messagebox.showerror(
                "Import Error",
                "Missing required columns:\n" + ", ".join(sorted(missing))
            )
            return

        # 4) ask dataset/display name
        display_name = simpledialog.askstring(
            "Import Excel",
            "Name this dataset (example: Jan 2026 Projects):"
        )
        if not display_name:
            return

        def clean(v):
            """Convert pandas NaN/NaT to None; trim strings."""
            if pd.isna(v):
                return None
            if isinstance(v, str):
                v = v.strip()
                return v if v != "" else None
            return v

        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()

        try:
            # 5) create new table name
            base = self._sanitize_table_name("project_" + display_name)
            table_name = self._unique_table_name(cur, base)

            # 6) create per-file table with SAME schema as tasks
            cur.execute(sql.SQL("""
                CREATE TABLE {tbl} (
                    ID INTEGER PRIMARY KEY,
                    CUSTOMER TEXT,
                    PROJECT_NAME TEXT NOT NULL,
                    DATE_OF_PO DATE NOT NULL,
                    LEADTIME REAL NOT NULL,
                    STATUS TEXT,
                    TARGET_DELIVERY DATE NULL,
                    DEPENDENCIES TEXT NULL,
                    LAMBDA_VAL REAL NOT NULL,
                    MU_VAL REAL NOT NULL,
                    NOS_VAL INTEGER NOT NULL,
                    WORK_TYPE TEXT NOT NULL,
                    REMARKS TEXT
                )
            """).format(tbl=sql.Identifier(table_name)))

            # 7) insert rows (auto-fix missing IDs)
            insert_q = sql.SQL("""
                INSERT INTO {tbl} (
                    ID, CUSTOMER, PROJECT_NAME, DATE_OF_PO, LEADTIME, STATUS,
                    TARGET_DELIVERY, DEPENDENCIES,
                    LAMBDA_VAL, MU_VAL, NOS_VAL,
                    WORK_TYPE, REMARKS
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (ID) DO UPDATE SET
                    CUSTOMER=EXCLUDED.CUSTOMER,
                    PROJECT_NAME=EXCLUDED.PROJECT_NAME,
                    DATE_OF_PO=EXCLUDED.DATE_OF_PO,
                    LEADTIME=EXCLUDED.LEADTIME,
                    STATUS=EXCLUDED.STATUS,
                    TARGET_DELIVERY=EXCLUDED.TARGET_DELIVERY,
                    DEPENDENCIES=EXCLUDED.DEPENDENCIES,
                    LAMBDA_VAL=EXCLUDED.LAMBDA_VAL,
                    MU_VAL=EXCLUDED.MU_VAL,
                    NOS_VAL=EXCLUDED.NOS_VAL,
                    WORK_TYPE=EXCLUDED.WORK_TYPE,
                    REMARKS=EXCLUDED.REMARKS
            """).format(tbl=sql.Identifier(table_name))

            # if IDs are blank, generate them
            next_id = 1
            # if some rows have IDs, make next_id larger than max
            if "ID" in df.columns:
                try:
                    max_id = pd.to_numeric(df["ID"], errors="coerce").max()
                    if pd.notna(max_id):
                        next_id = int(max_id) + 1
                except Exception:
                    pass

            for _, row in df.iterrows():
                rid = clean(row.get("ID"))

                # auto-generate if missing
                if rid is None:
                    rid = next_id
                    next_id += 1
                else:
                    rid = int(float(rid))  # handles "1.0" from Excel

                customer = clean(row.get("CUSTOMER"))
                pname = clean(row.get("PROJECT NAME"))
                date_po = clean(row.get("DATE OF PO"))
                lead = clean(row.get("LEADTIME"))

                # enforce required values
                if pname is None or date_po is None or lead is None:
                    continue

                cur.execute(insert_q, (
                    rid,
                    customer,
                    pname,
                    date_po,
                    float(lead),
                    clean(row.get("STATUS")),
                    clean(row.get("TARGET DELIVERY")),
                    clean(row.get("DEPENDENCIES")),
                    float(clean(row.get("LAMBDA"))) if clean(row.get("LAMBDA")) is not None else 0.0,
                    float(clean(row.get("MU"))) if clean(row.get("MU")) is not None else 0.0,
                    int(float(clean(row.get("NO. OF SERVER")))) if clean(row.get("NO. OF SERVER")) is not None else 1,
                    clean(row.get("WORK TYPE")),
                    clean(row.get("REMARKS"))
                ))

            # 8) register in imported_files
            cur.execute("""
                INSERT INTO imported_files (display_name, table_name, file_path)
                VALUES (%s, %s, %s)
            """, (display_name, table_name, path))

            conn.commit()

        except Exception as e:
            conn.rollback()
            messagebox.showerror("Import Error", str(e))
            return

        finally:
            cur.close()
            conn.close()

        # refresh dropdown + auto-select dataset
        if hasattr(self, "refresh_dataset_list"):
            self.refresh_dataset_list()

        self.current_tasks_table = table_name
        self.load_tasks()

        messagebox.showinfo("Success", f"Imported '{display_name}' into table: {table_name}")

    def refresh_dataset_list(self):
        conn = psycopg2.connect(**DB_PARAMS)
        cur = conn.cursor()

        # get all saved datasets
        cur.execute("SELECT display_name, table_name FROM imported_files ORDER BY imported_at DESC")
        rows = cur.fetchall()

        valid = []
        for display, table in rows:
            # check if the table still exists
            cur.execute("""
                SELECT to_regclass(%s)
            """, (table,))
            exists = cur.fetchone()[0] is not None

            if exists:
                valid.append((display, table))
            else:
                # remove dead reference from imported_files
                cur.execute("DELETE FROM imported_files WHERE table_name = %s", (table,))

        conn.commit()
        cur.close()
        conn.close()

        self.dataset_map = {"Default (tasks)": "tasks"}
        for display, table in rows:
            self.dataset_map[str(display)] = str(table)

        values = list(self.dataset_map.keys())

        if hasattr(self, "dataset_combo_one"):
            self.dataset_combo_one["values"] = values
        if hasattr(self, "dataset_combo_two"):
            self.dataset_combo_two["values"] = values

        if self.dataset_var.get() not in self.dataset_map:
            self.dataset_var.set("Default (tasks)")

    def refresh_sidebar(self):
        self.refresh_user_session()
        self.refresh_key_metrics()
        self.refresh_activity_log()

    def _safe_filename(self, s: str) -> str:
        bad = r'<>:"/\|?*'
        s = "".join("_" if ch in bad else ch for ch in str(s))
        s = s.strip().strip(".")
        return s or "dataset"

    def _dataset_prefix_and_date(self) -> str:
        display = getattr(self, "dataset_var", None)
        ds_name = display.get() if display else "Default (tasks)"
        ds_name = self._safe_filename(ds_name.replace(" ", "_"))

        today = date.today().isoformat()   # e.g. 2026-01-24
        return f"{ds_name}_{today}"

    def export_tasks_excel(self):
        prefix = self._dataset_prefix_and_date()
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{prefix}_All_Projects_Tasks.xlsx"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "All Projects"

        self._excel_write_tree(ws, "ALL PROJECTS", self.task_tree, start_row=1)
        ws.freeze_panes = "A2"

        wb.save(path)
        messagebox.showinfo("Success", "Exported ALL PROJECTS (Tasks tab).")

    def export_cpm_excel(self):
        prefix = self._dataset_prefix_and_date()
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{prefix}_CPM_Report.xlsx"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "CPM"

        row = 1
        row, cpm_end_col, cpm_section_top = self._excel_write_tree(ws, "CPM RESULTS", self.cpm_tree, start_row=row)
        row += 3

        if hasattr(self, "cpm_all_tasks_tree"):
            row, _, _ = self._excel_write_tree(ws, "ALL PROJECTS", self.cpm_all_tasks_tree, start_row=row)
            row += 3

        if hasattr(self, "cpm_gantt_fig"):
            self._excel_insert_fig_right(ws, self.cpm_gantt_fig, cpm_end_col, cpm_section_top, width_px=850, height_px=350)

        ws.freeze_panes = "A2"
        wb.save(path)
        messagebox.showinfo("Success", "Exported CPM report.")

    def export_queueing_excel(self):
        prefix = self._dataset_prefix_and_date()
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{prefix}_Queueing_Report.xlsx"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Queueing"

        row = 1
        row, _, _ = self._excel_write_tree(ws, "SUMMARY BY WORK TYPE", self.queue_summary_tree, start_row=row)
        row += 3

        row, q_end_col, q_section_top = self._excel_write_tree(ws, "QUEUEING - ALL PROJECTS", self.queue_tree, start_row=row)
        row += 3

        if hasattr(self, "queue_heatmap_fig"):
            self._excel_insert_fig_right(ws, self.queue_heatmap_fig, q_end_col, q_section_top, width_px=850, height_px=300)

        ws.freeze_panes = "A2"
        wb.save(path)
        messagebox.showinfo("Success", "Exported Queueing report.")


    def _excel_insert_fig_right(self, ws, fig, table_end_col, top_row, width_px=720, height_px=280):
        # save fig to temp png
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        tmp.close()
        fig.savefig(tmp.name, dpi=150, bbox_inches="tight")

        img = XLImage(tmp.name)
        img.width = width_px
        img.height = height_px

        anchor_col = table_end_col + 2
        anchor_cell = f"{get_column_letter(anchor_col)}{top_row}"
        ws.add_image(img, anchor_cell)


    def _excel_write_tree(self, ws, title, tree, start_row):
        # --- fills (same as yours) ---
        FILL_GREEN  = PatternFill(fill_type="solid", fgColor="FFC6EFCE")
        FILL_RED    = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
        FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFFFEB9C")
        FILL_PURPLE = PatternFill(fill_type="solid", fgColor="FFE6D9FF")
        FILL_CYAN   = PatternFill(fill_type="solid", fgColor="FFB2EBF2")

        # Styles
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        title_font = Font(bold=True, size=14)
        title_fill = PatternFill("solid", fgColor="D9E1F2")  # light blue
        title_alignment = Alignment(horizontal="left", vertical="center")

        MIN_WIDTH = 14
        MAX_WIDTH = 45
        PADDING = 3

        cols = list(tree["columns"])

        headers = []
        for c in cols:
            h = tree.heading(c).get("text")
            headers.append(h if h else c)

        r = start_row
        section_start_row = r

        # ---- Section title ----
        last_col = len(headers) if headers else 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        r += 1

        # ---- Header row ----
        header_row = r
        for j, header in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=j, value=header)
            c.font = header_font
            c.alignment = header_alignment
            c.border = border
        ws.row_dimensions[header_row].height = 22
        r += 1

        # ---- Data rows ----
        for item_id in tree.get_children():
            values = tree.item(item_id, "values")
            row_fill = None

            # TASKS TAB coloring
            if tree is getattr(self, "task_tree", None):
                status = str(values[5]).upper()  # STATUS index in tasks tree
                if status in ("ONGOING", "ON TIME"):
                    row_fill = FILL_GREEN
                elif status == "DELAY":
                    row_fill = FILL_RED
                elif status == "ON-HOLD":
                    row_fill = FILL_YELLOW
                elif "SUBJECT" in status:
                    row_fill = FILL_PURPLE
                elif "COMPLETED" in status:
                    row_fill = FILL_CYAN

            # CPM RESULTS coloring
            elif tree is getattr(self, "cpm_tree", None):
                critical = str(values[-1]).upper()
                if critical == "YES":
                    row_fill = FILL_RED

            # QUEUEING ALL PROJECTS coloring (rho)
            elif tree is getattr(self, "queue_tree", None):
                try:
                    rho = float(values[7])  # RHO column index
                    if rho >= 1.0:
                        row_fill = FILL_RED
                    elif rho >= 0.85:
                        row_fill = FILL_YELLOW
                except:
                    row_fill = None

            # SUMMARY coloring (FLAG)
            elif tree is getattr(self, "queue_summary_tree", None):
                # If your FLAG column is last:
                flag = str(values[-1]).upper()
                if "UNSTABLE" in flag or "INVALID" in flag:
                    row_fill = FILL_RED
                elif "WARNING" in flag:
                    row_fill = FILL_YELLOW
                elif "OK" in flag:
                    row_fill = FILL_GREEN

            for j, v in enumerate(values, start=1):
                cell = ws.cell(row=r, column=j, value=v)
                if row_fill:
                    cell.fill = row_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            r += 1

        # ---- Column widths ----
        for j, header in enumerate(headers, start=1):
            col_letter = get_column_letter(j)
            max_len = len(str(header))

            for rr in range(header_row, r):
                val = ws.cell(row=rr, column=j).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))

            width = max_len + PADDING
            width = max(MIN_WIDTH, min(width, MAX_WIDTH))
            ws.column_dimensions[col_letter].width = width

        end_col = last_col
        return r, end_col, section_start_row
    
    def logout_user(self):
        if not messagebox.askyesno("Logout", "Are you sure you want to log out?"):
            return

        # ---- Clear session data ----
        self.current_user = ""
        self.current_role = ""
        self.current_tasks_table = "tasks"
        self.dataset_var.set("Default (tasks)")

        # ---- Clear sidebar UI ----
        self.lbl_user.config(text="Logged in: --")
        self.lbl_role.config(text="Role: --")
        self.last_refresh_lbl.config(text="Last refresh: --")

        self.lbl_avg_lead.config(text="Avg Lead Time: -- days")
        self.lbl_on_time.config(text="On-Time Rate: --%")
        self.lbl_avg_delay.config(text="Avg Delay: -- days")
        self.lbl_critical.config(text="Critical Projects: --")

        self.activity_list.delete(0, "end")
        self.activity_list.insert("end", "No recent activity")

        # ---- Clear main dashboard KPIs ----
        for var in self.kpi_vars.values():
            var.set("0")

        # ---- Redirect to Login page ----
        self.controller.show_frame(LoginPage)
# ---------------- RUN ----------------
if __name__=="__main__":
    create_db()
    app = App()
    app.mainloop()
